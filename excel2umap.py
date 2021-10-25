from sklearn.datasets import load_digits
import matplotlib.pyplot as plt
import umap
import numpy as np
from openpyxl import load_workbook
import time
import yaml
import sys
import os

def excel2umap(excel_path, sheet2label, save_result_image_path):
    wb = load_workbook(filename=excel_path, read_only=True)
    x = []
    y = []
    for sheetname in sheet2label.keys():
        ws = wb[sheetname]
        for i in range(2, ws.max_row+1):
            row = []
            for j in range(2, ws.max_column+1):
                v = ws.cell(row=i, column=j).value
                row.append(v)
            x.append(row)
            y.append(sheet2label[sheetname])
    x = np.array(x)
    y = np.array(y)

    reducer = umap.UMAP(random_state=42)
    embedding = reducer.fit_transform(x)

    plt.scatter(embedding[:, 0], embedding[:, 1], c=y, cmap='Spectral')
    plt.gca().set_aspect('equal', 'datalim')
    plt.colorbar(boundaries=np.arange(1, max(sheet2label.values())+2)-0.5).set_ticks(np.arange(1, max(sheet2label.values())+1))
    plt.title(f'UMAP projection of the {len(sheet2label.keys())} cohorts image statistics')
    os.makedirs(os.path.dirname(save_result_image_path), exist_ok=True)
    plt.savefig(save_result_image_path)

if __name__ == '__main__':
    start = time.time()
    yaml_path = sys.argv[1]
    with open(yaml_path, 'r') as f:
        param_sets = yaml.safe_load(f)
    for set_name, param in param_sets.items():
        print(set_name)
        print(param)
        excel_path = param.get('excel_path')
        sheet2label = param.get('sheet2label')
        save_result_image_path = param.get('save_result_image_path')
        excel2umap(excel_path, sheet2label, save_result_image_path)
    end = time.time()
    print('done! time:', (end-start)/60)

