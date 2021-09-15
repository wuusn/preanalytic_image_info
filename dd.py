from sklearn.datasets import load_digits
import matplotlib.pyplot as plt
import umap
import numpy as np
from openpyxl import load_workbook
import time
start = time.time()

#excel_path = '/mnt/md0/_datasets/OralCavity/im_state/Oral_Multi_Site_Quality_3.xlsx'
#wb = load_workbook(filename=excel_path, read_only=True)
#x = []
#y = []
#for d in [1,2,3,4,5]:
#    ws = wb[f'D{d}']
#    for i in range(2, ws.max_row+1):
#        row = []
#        for j in range(2, ws.max_column+1):
#            v = ws.cell(row=i, column=j).value
#            row.append(v)
#        x.append(row)
#        y.append(d)
#x = np.array(x)
#y = np.array(y)

#print(x.shape,y.shape)
import pickle
with open(f'x_y.pkl', 'rb') as f:
    x,y = pickle.load(f)

#digits = load_digits()
#print(digits.target[0])
#print(digits.target.__class__)
#print(digits.target.shape)
#reducer = umap.UMAP(random_state=42)
#embedding = reducer.fit_transform(x)
with open(f'embedding.pkl', 'rb') as f:
    embedding = pickle.load(f)
#
labels = ('D1', 'D2', 'D3', 'D4', 'D5')
plt.scatter(embedding[:, 0], embedding[:, 1], c=y, cmap='Spectral', s=5)
plt.gca().set_aspect('equal', 'datalim')
cbar = plt.colorbar(boundaries=np.arange(1,7)-0.5, extendfrac='auto')
cbar.set_ticks(np.array([1,2,3,4,5]))
cbar.set_ticklabels(labels)
plt.title('UMAP projection of the five cohorts image statistics')
plt.savefig('umap.png')
end = time.time()
print('time:', (end-start)/60)
