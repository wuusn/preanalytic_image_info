from sklearn.datasets import load_digits
import matplotlib.pyplot as plt
import umap
import numpy as np
from openpyxl import load_workbook

excel_path = '/mnt/md0/_datasets/OralCavity/im_state/Oral_Multi_Site_Quality_3.xlsx'
wb = load_workbook(filename=excel_path, read_only=True)
x = []
y = []
for d in [1,2,3,4,5]:
    ws = wb[f'D{d}']
    for i in range(2, ws.max_row+1):
        row = []
        for j in range(2, ws.max_column+1):
            v = ws.cell(row=i, column=j).value
            row.append(v)
        x.append(row)
        y.append(d)
x = np.array(x)
y = np.array(y)

print(x.shape,y.shape)


#digits = load_digits()
#print(digits.target[0])
#print(digits.target.__class__)
#print(digits.target.shape)
#reducer = umap.UMAP(random_state=42)
#embedding = reducer.fit_transform(x)
##
#plt.scatter(embedding[:, 0], embedding[:, 1], c=y, cmap='Spectral', s=5)
#plt.gca().set_aspect('equal', 'datalim')
#plt.colorbar(boundaries=np.arange(11)-0.5).set_ticks(np.arange(10))
#plt.title('UMAP projection of the five cohorts image statistics')
#plt.savefig('umap.png')
